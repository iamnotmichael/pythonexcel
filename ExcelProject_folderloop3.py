

# Import Packages
import openpyxl, os, re

# Check and change directory if necessary
os.getcwd()
os.chdir('C:/Users/Michael/OneDrive/EXCEL PYTHON/FILES')
os.getcwd()
folderpath = 'C:/Users/Michael/OneDrive/EXCEL PYTHON/FILES/mass'

# Create variable with list of all files in above folder
from os import listdir
from os.path import isfile, join
folderfiles = [f for f in listdir(folderpath) if isfile(join(folderpath, f))]

# Find Report Date Nested in Text String at Bottom of Excel Sheet
def finddate(sheet):
    lastRowText = sheet.cell(row = sheet.max_row, column = 1)
    matchgroup = re.compile(r'\d{2}/\d{2}/\d{4}')
    matchResult = matchgroup.search(lastRowText.value)
    exceldate = matchResult.group(0)
    return exceldate

# Get Row Locations for Categories and create Dict with List Values of row locations
def getlocations(sheet):
    locationDict = {'Credit.Apps':[],'Credit.Operators':[],'Credit.Stips.Zone.1':[],'Credit.Stips.Zone.2':[],'Credit.Stips.Zone.4':[]}
    for i in range (1,sheet.max_row-3):
        g = sheet.cell(row = i, column = 1)   
        if g.value == 'Credit.Apps':
            locationDict['Credit.Apps'].append(g.row)
        elif g.value == 'Credit.Operators':
            locationDict['Credit.Operators'].append(g.row)
        elif g.value == 'Credit.Stips.Zone.1':
            locationDict['Credit.Stips.Zone.1'].append(g.row)       
        elif g.value == 'Credit.Stips.Zone.2':
            locationDict['Credit.Stips.Zone.2'].append(g.row)     
        elif g.value == 'Credit.Stips.Zone.4':
            locationDict['Credit.Stips.Zone.4'].append(g.row)
    return locationDict

# Read and Copy Excel Data
def copydata(sheet):
    exceldate = finddate(sheet)
    locationDict = getlocations(sheet)
    for k in locationDict:
        if k == 'Credit.Apps' or k == 'Credit.Operators':
            for j in range(getlocations(sheet)[k][0],getlocations(sheet)[k][1]):
                appsmax = sheetAPPS.max_row+1
                for i in range(1,16):
                    if i == 1:
                        appwrite = sheetAPPS.cell(row=appsmax, column = i)
                        appwrite.value = exceldate
                    else:
                        e = sheet.cell(row=j,column=i)
                        appwrite = sheetAPPS.cell(row=appsmax, column = i)
                        appwrite.value = e.value
        else:
            for j in range(getlocations(sheet)[k][0],getlocations(sheet)[k][1]):
                stipsmax = sheetSTIPS.max_row+1
                for i in range(1,16):
                    if i == 1:
                        appwrite = sheetSTIPS.cell(row=stipsmax, column = i)
                        appwrite.value = exceldate
                    else:
                        e = sheet.cell(row=j,column=i)
                        appwrite = sheetSTIPS.cell(row=stipsmax, column = i)
                        appwrite.value = e.value
                
    dateappmax = sheetTrackerApp.max_row+1
    appwrite = sheetTrackerApp.cell(row=dateappmax, column = 1)
    appwrite.value = exceldate
    datestipsmax = sheetTrackerStip.max_row+1
    stipwrite = sheetTrackerStip.cell(row=datestipsmax, column = 1)
    stipwrite.value = exceldate

# Function that contains other functions to actually copy the data and save the files
def copyfunction(sheet):
    reportdate = finddate(sheet)
    for j in range(1,sheetTrackerApp.max_row+1):
        e = sheetTrackerApp.cell(row = j, column = 1)
        if reportdate == e.value:
            return
    locations = getlocations(sheet)
    runall = copydata(sheet)
    wbapps.save('APPS PHONE.xlsx')
    wbstips.save('STIPS PHONE.xlsx')
    wbtracker.save('DATETRACKER.xlsx')

wbapps = openpyxl.load_workbook('APPS PHONE.xlsx')
wbstips = openpyxl.load_workbook('STIPS PHONE.xlsx')
wbtracker = openpyxl.load_workbook('DATETRACKER.xlsx') 
sheetAPPS = wbapps.get_sheet_by_name('Phones')
sheetSTIPS = wbstips.get_sheet_by_name('Phones')
sheetTrackerApp = wbtracker.get_sheet_by_name('APP')
sheetTrackerStip = wbtracker.get_sheet_by_name('STIPS')

while True:     
    for i in range(len(folderfiles)):
        folderfileName = "mass/"+folderfiles[i]
        wbexcel = openpyxl.load_workbook(folderfileName)
        sheet = wbexcel.get_sheet_by_name('Originations Phone Report')
        copyfunction(sheet)
    break
